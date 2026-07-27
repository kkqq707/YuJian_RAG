"""文件上传安全校验模块

支持格式: .txt / .md / .pdf / .docx / .xlsx
校验: 扩展名、MIME 类型、文件大小、文件签名、文件名安全、空文件、PDF 加密

安全策略:
- 禁止可执行文件、脚本、压缩包、宏文件
- 禁止路径穿越
- 文件名安全化处理
"""

from __future__ import annotations

import hashlib
import logging
import mimetypes
import re
from pathlib import Path
from typing import Optional

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# 常量
# ---------------------------------------------------------------------------

ALLOWED_EXTENSIONS = {".txt", ".md", ".pdf", ".docx", ".xlsx"}
FORBIDDEN_EXTENSIONS = {
    ".exe", ".bat", ".cmd", ".ps1", ".vbs", ".wsf", ".msi", ".scr",
    ".py", ".pyc", ".pyd", ".pyo", ".pyw",
    ".js", ".jsx", ".ts", ".tsx", ".vbs",
    ".html", ".htm", ".xhtml", ".shtml",
    ".zip", ".rar", ".7z", ".tar", ".gz", ".bz2", ".xz", ".lz", ".lzma",
    ".docm", ".xlsm", ".pptm",  # 宏文件
    ".dll", ".so", ".dylib",
    ".sh", ".bash", ".zsh", ".fish", ".csh",
    ".jar", ".war", ".ear",
    ".php", ".asp", ".aspx", ".jsp",
    ".class", ".swf",
}
MAX_FILENAME_LENGTH = 150
MAX_FILE_SIZE_MB = 20
MAX_FILE_SIZE_BYTES = MAX_FILE_SIZE_MB * 1024 * 1024
MAX_FILES_PER_UPLOAD = 10
SAFE_FILENAME_PATTERN = re.compile(r"[^a-zA-Z0-9_\-.一-鿿\(\)（） ]")

# 常见文件签名（魔数）
FILE_SIGNATURES = {
    b"\x25\x50\x44\x46": ".pdf",   # PDF
    b"\x50\x4b\x03\x04": ".docx",  # DOCX / XLSX (both are Office Open XML)
}

# ---------------------------------------------------------------------------
# 公开函数
# ---------------------------------------------------------------------------


def validate_file_extension(filename: str) -> Path:
    """校验文件扩展名是否在允许列表中。

    Parameters
    ----------
    filename : str
        原始文件名

    Returns
    -------
    Path
        安全化后的文件名 Path 对象

    Raises
    ------
    ValueError
        扩展名不允许或文件名不安全
    """
    if not filename or not filename.strip():
        raise ValueError("文件名为空")

    name = filename.strip()
    path = Path(name)
    suffix = path.suffix.lower()

    if suffix in FORBIDDEN_EXTENSIONS:
        raise ValueError(f"禁止上传的文件类型: {suffix}")

    if suffix not in ALLOWED_EXTENSIONS:
        raise ValueError(
            f"不支持的文件类型 '{suffix}'，当前支持: "
            f"{', '.join(sorted(ALLOWED_EXTENSIONS))}"
        )

    return path


def validate_filename_safety(filename: str) -> str:
    """验证文件名安全性，禁止路径穿越和非法字符。

    Parameters
    ----------
    filename : str
        原始文件名

    Returns
    -------
    str
        清理后的安全文件名

    Raises
    ------
    ValueError
        文件名不安全
    """
    if not filename or not filename.strip():
        raise ValueError("文件名为空")

    name = filename.strip()

    # 禁止路径穿越
    if ".." in name:
        raise ValueError("文件名包含非法字符 '..'，禁止路径穿越")

    if "/" in name or "\\" in name:
        raise ValueError("文件名包含路径分隔符，禁止路径穿越")

    # 检查绝对路径
    if name.startswith("/") or name.startswith("\\"):
        raise ValueError("文件名不能以路径分隔符开头")

    # Windows 盘符检查
    if len(name) >= 2 and name[1] == ":":
        raise ValueError("文件名不能包含盘符")

    # 文件名长度限制
    if len(name) > MAX_FILENAME_LENGTH:
        raise ValueError(
            f"文件名过长 ({len(name)} 字符)，最大允许 {MAX_FILENAME_LENGTH} 字符"
        )

    return name


def sanitize_filename(filename: str) -> str:
    """将文件名转换为安全的存储名称。

    保留中文、英文、数字、下划线、连字符、点号和括号。
    其他字符替换为下划线。

    Parameters
    ----------
    filename : str
        原始文件名

    Returns
    -------
    str
        安全化后的文件名
    """
    name = Path(filename).stem
    suffix = Path(filename).suffix.lower()
    safe_name = SAFE_FILENAME_PATTERN.sub("_", name)
    # 合并连续下划线
    safe_name = re.sub(r"_+", "_", safe_name).strip("_")
    if not safe_name:
        safe_name = "unnamed"
    # 截断以保证加上后缀后不超过长度限制
    max_name_len = MAX_FILENAME_LENGTH - len(suffix)
    if len(safe_name) > max_name_len:
        safe_name = safe_name[:max_name_len]
    return safe_name + suffix


def validate_file_size(file_size: int) -> None:
    """校验文件大小是否在允许范围内。

    Parameters
    ----------
    file_size : int
        文件大小（字节）

    Raises
    ------
    ValueError
        文件为空或超大
    """
    if file_size <= 0:
        raise ValueError("文件为空，禁止上传空文件")
    if file_size > MAX_FILE_SIZE_BYTES:
        size_mb = file_size / (1024 * 1024)
        raise ValueError(
            f"文件过大 ({size_mb:.1f} MB)，最大允许 {MAX_FILE_SIZE_MB} MB"
        )


def validate_file_signature(file_content: bytes, expected_suffix: str) -> None:
    """验证文件签名（魔数）是否与扩展名匹配。

    仅对已知签名的文件类型进行检测；签名未知则不报错。

    Parameters
    ----------
    file_content : bytes
        文件内容的前几个字节
    expected_suffix : str
        期望的文件扩展名

    Raises
    ------
    ValueError
        签名不匹配
    """
    if len(file_content) < 4:
        return  # 太短无法检测

    for signature, suffix in FILE_SIGNATURES.items():
        if expected_suffix == suffix:
            if not file_content.startswith(signature):
                raise ValueError(
                    f"文件签名不匹配：扩展名为 {expected_suffix}，"
                    f"但文件内容不符合该格式"
                )


def detect_mime_type(file_content: bytes) -> str:
    """探测文件的 MIME 类型。

    Parameters
    ----------
    file_content : bytes
        文件内容

    Returns
    -------
    str
        MIME 类型字符串
    """
    import tempfile
    # mimetypes 需要文件路径，我们写临时文件来检测
    # 对于大多数情况，基于文件头的简单判断即可
    if file_content.startswith(b"\x25\x50\x44\x46"):
        return "application/pdf"
    if file_content.startswith(b"\x50\x4b\x03\x04"):
        return "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    # 对于文本文件，尝试解码（使用 errors='replace' 避免截断多字节字符）
    try:
        text = file_content[:1024].decode("utf-8", errors="replace")
        # 如果替换字符过多说明不是文本文件
        if text.count("�") < len(text) * 0.1:
            return "text/plain"
    except Exception:
        pass
    try:
        text = file_content[:1024].decode("gb18030", errors="replace")
        if text.count("�") < len(text) * 0.1:
            return "text/plain"
    except Exception:
        pass
    return "application/octet-stream"


def validate_mime_type(file_content: bytes, expected_suffix: str) -> None:
    """验证 MIME 类型与扩展名是否大体匹配。

    Parameters
    ----------
    file_content : bytes
        文件内容
    expected_suffix : str
        期望的文件扩展名

    Raises
    ------
    ValueError
        MIME 类型明显不匹配
    """
    mime = detect_mime_type(file_content)
    suffix_mime_map = {
        ".txt": ("text/",),
        ".md": ("text/",),
        ".pdf": ("application/pdf",),
        ".docx": ("application/vnd.openxmlformats-officedocument.wordprocessingml",),
        ".xlsx": ("application/vnd.openxmlformats-officedocument.spreadsheetml",),
    }
    expected_prefixes = suffix_mime_map.get(expected_suffix)
    if expected_prefixes is None:
        return  # 未知类型不校验
    if not any(mime.startswith(prefix) for prefix in expected_prefixes):
        raise ValueError(
            f"MIME 类型不匹配：扩展名为 {expected_suffix}，"
            f"检测到 MIME 类型为 {mime}"
        )


def is_pdf_encrypted(file_content: bytes) -> bool:
    """检测 PDF 是否加密。

    通过检查 PDF 头部的 /Encrypt 标记。

    Parameters
    ----------
    file_content : bytes
        PDF 文件内容

    Returns
    -------
    bool
        是否加密
    """
    try:
        text = file_content.decode("latin-1", errors="replace")
        # 简单检测：/Encrypt 标记
        if "/Encrypt" in text:
            return True
        # 也检查是否需要密码
        if "/EncryptMetadata" in text:
            return True
    except Exception:
        pass
    return False


def can_parse_document(file_path: Path) -> bool:
    """检测文档是否可以正常解析。

    Parameters
    ----------
    file_path : Path
        文件路径

    Returns
    -------
    bool
        是否可以正常解析

    Raises
    ------
    ValueError
        文档无法解析时抛出
    """
    suffix = file_path.suffix.lower()

    if suffix == ".pdf":
        try:
            from pypdf import PdfReader
            reader = PdfReader(str(file_path))
            if len(reader.pages) == 0:
                raise ValueError("PDF 文件没有可读页面")
            # 尝试读取第一页
            try:
                reader.pages[0].extract_text()
            except Exception:
                pass  # 某些页面可能没有文本，不报错
        except Exception as e:
            if "encrypted" in str(e).lower() or "password" in str(e).lower():
                raise ValueError("PDF 文件已加密或需要密码，无法处理")
            raise ValueError(f"PDF 文件解析失败: {_safe_str(e)}")

    elif suffix == ".docx":
        try:
            import docx as _docx
            doc = _docx.Document(str(file_path))
            paragraphs = [p.text for p in doc.paragraphs if p.text and p.text.strip()]
            if not paragraphs:
                raise ValueError("DOCX 文件中未找到可读文本")
        except Exception as e:
            raise ValueError(f"DOCX 文件解析失败: {_safe_str(e)}")

    elif suffix in (".txt", ".md"):
        # 尝试用不同编码读取
        raw = file_path.read_bytes()
        if len(raw.strip()) == 0:
            raise ValueError("文本文件内容为空")

    elif suffix == ".xlsx":
        try:
            import openpyxl
            wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
            if not wb.sheetnames:
                raise ValueError("XLSX 文件中未找到工作表")
            # Check at least one sheet has data
            has_data = False
            for name in wb.sheetnames:
                sheet = wb[name]
                if sheet.max_row and sheet.max_row > 0:
                    has_data = True
                    break
            wb.close()
            if not has_data:
                raise ValueError("XLSX 文件中未找到可读数据")
        except Exception as e:
            raise ValueError(f"XLSX 文件解析失败: {_safe_str(e)}")

    return True


def compute_file_hash(file_path: Path) -> str:
    """计算文件的 SHA-256 哈希值。

    Parameters
    ----------
    file_path : Path
        文件路径

    Returns
    -------
    str
        十六进制哈希字符串
    """
    sha256 = hashlib.sha256()
    with open(file_path, "rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            sha256.update(chunk)
    return sha256.hexdigest()


def compute_bytes_hash(data: bytes) -> str:
    """计算字节数据的 SHA-256 哈希值。

    Parameters
    ----------
    data : bytes

    Returns
    -------
    str
    """
    return hashlib.sha256(data).hexdigest()


def validate_upload(
    filename: str,
    file_content: bytes,
    file_size: int,
) -> dict:
    """执行完整的上传文件安全校验。

    校验顺序:
    1. 文件名安全性
    2. 文件扩展名
    3. 文件大小
    4. 文件签名
    5. MIME 类型
    6. PDF 加密检测

    Parameters
    ----------
    filename : str
        原始文件名
    file_content : bytes
        文件完整内容
    file_size : int
        文件大小（字节）

    Returns
    -------
    dict
        {
            "valid": bool,
            "error": str | None,
            "original_name": str,
            "safe_filename": str,
            "file_hash": str,
            "file_type": str,
        }
    """
    result = {
        "valid": False,
        "error": None,
        "original_name": filename,
        "safe_filename": "",
        "file_hash": "",
        "file_type": "",
    }

    try:
        # 1. 文件名安全性
        safe_name = validate_filename_safety(filename)

        # 2. 扩展名校验
        path = validate_file_extension(safe_name)
        suffix = path.suffix.lower()
        result["file_type"] = suffix

        # 3. 文件大小
        validate_file_size(file_size)

        # 4. 文件签名
        validate_file_signature(file_content, suffix)

        # 5. MIME 类型
        validate_mime_type(file_content, suffix)

        # 6. PDF 加密检测
        if suffix == ".pdf":
            if is_pdf_encrypted(file_content):
                raise ValueError("PDF 文件已加密，无法处理")

        # 7. 计算哈希
        file_hash = compute_bytes_hash(file_content)
        result["file_hash"] = file_hash

        # 8. 安全化文件名
        result["safe_filename"] = sanitize_filename(safe_name)

        result["valid"] = True
        return result

    except ValueError as e:
        result["error"] = str(e)
        return result


def generate_stored_name(upload_id: str, safe_filename: str) -> str:
    """生成存储文件名: {upload_id}_{safe_filename}。

    Parameters
    ----------
    upload_id : str
        上传唯一 ID
    safe_filename : str
        已安全化的文件名

    Returns
    -------
    str
    """
    return f"{upload_id}_{safe_filename}"


# ---------------------------------------------------------------------------
# 内部辅助
# ---------------------------------------------------------------------------


def _safe_str(exc: Exception) -> str:
    """从异常中提取安全的一行摘要。"""
    return str(exc).split("\n")[0][:150]
